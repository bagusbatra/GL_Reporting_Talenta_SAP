"""
GL Reporting Python Service - SURABAYA (Port 8092)
====================================================
Service untuk fetch data Talenta + generate Excel SAP format.
Support 4 extraction strategy: B, D, E, F

Endpoint:
    POST /run    -> jalankan extraction
    GET  /health -> health check

Dipanggil oleh Laravel via HTTP request.
"""

import hmac
import hashlib
import base64
import json
import os
import time
import traceback
from datetime import datetime
from urllib.parse import urlparse

import requests
import pandas as pd
from flask import Flask, request, jsonify
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CONFIG
# ============================================================
SERVICE_PORT = 8092
SERVICE_NAME = "GL Reporting Service - SURABAYA"
OUTPUT_BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'storage', 'app', 'gl_outputs')

SAP_BASE_SETTINGS = {
    'doc_type': 'KA',
    'currency': 'IDR',
    'reference': 'TALENTA',
}

MONTH_NAMES = ["", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
               "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]

# ============================================================
# TALENTA API CLIENT (sama dengan Semarang)
# ============================================================
def api_request(url, client_id, client_secret):
    try:
        parsed = urlparse(url)
        request_line = f"GET {parsed.path}?{parsed.query} HTTP/1.1"
        date_string = datetime.utcnow().strftime('%a, %d %b %Y %H:%M:%S GMT')
        signing_string = f"date: {date_string}\n{request_line}"
        digest = hmac.new(client_secret.encode(), signing_string.encode(), hashlib.sha256).digest()
        signature = base64.b64encode(digest).decode()
        headers = {
            "Authorization": f'hmac username="{client_id}", algorithm="hmac-sha256", headers="date request-line", signature="{signature}"',
            "Date": date_string
        }
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 429:
            time.sleep(2)
            return api_request(url, client_id, client_secret)
        else:
            return None
    except Exception as e:
        print(f"API request error: {e}")
        return None


def get_ledger_report_history(year, month, client_id, client_secret):
    url = f"https://api.mekari.com/v2/talenta/v3/ledger/report/history?year={year}&month={month}&page=1&limit=100"
    data = api_request(url, client_id, client_secret)
    if data and 'data' in data:
        return data['data'].get('ledger_results', [])
    return []


def filter_reports_by_ledger_code(reports, target_ledger_code):
    matching = []
    for r in reports:
        code = r.get('ledger_code', '').strip()
        name = r.get('ledger_name', '').strip()
        if (code == target_ledger_code or name == target_ledger_code or
                target_ledger_code in code or target_ledger_code in name):
            matching.append(r)
    if not matching:
        return None
    matching.sort(key=lambda r: r.get('created_at', ''), reverse=True)
    return matching[0]


def get_report_detail(report_id, ledger_id, client_id, client_secret):
    url = f"https://api.mekari.com/v2/talenta/v3/ledger/report/{report_id}?ledger_id={ledger_id}"
    data = api_request(url, client_id, client_secret)
    if data and 'data' in data:
        return data['data'].get('ledger_result', {})
    return {}


def parse_amount(val):
    if isinstance(val, str):
        try:
            return float(val.replace(',', ''))
        except:
            return 0
    return val or 0


# ============================================================
# HELPER MATCHERS
# ============================================================
def match_by_account_number_and_keywords(mappings, account_number, account_name_lower):
    """Match by account_number, lalu kalau multi-variant pakai keyword priority"""
    candidates = [m for m in mappings if m['account_number'] == account_number]
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]['mapping_key']

    # Multi-variant: pilih keyword paling spesifik
    best_match = None
    best_kw_len = 0
    for m in candidates:
        keywords = m.get('match_keywords') or []
        for kw in keywords:
            if kw.lower() in account_name_lower and len(kw) > best_kw_len:
                best_match = m['mapping_key']
                best_kw_len = len(kw)
    return best_match if best_match else candidates[0]['mapping_key']


def match_by_match_account_name(mappings, account_number, account_name_lower):
    """
    Match khusus untuk Strategy B Surabaya (gl_driver.py):
    Setiap 2010000005 datang sebagai report_item terpisah dengan account_name berbeda.
    Match pakai field match_account_name.
    """
    candidates = [m for m in mappings if m['account_number'] == account_number]
    if not candidates:
        return None
    if len(candidates) == 1 and not candidates[0].get('match_account_name'):
        return candidates[0]['mapping_key']

    # Cari yang match_account_name-nya cocok
    for m in candidates:
        match_name = (m.get('match_account_name') or '').lower()
        if match_name and (match_name == account_name_lower or match_name in account_name_lower):
            return m['mapping_key']

    # Fallback: keyword match
    return match_by_account_number_and_keywords(mappings, account_number, account_name_lower)


# ============================================================
# STRATEGY B: Hybrid Cost Center + Aggregate (Surabaya variant)
# ============================================================
def extract_strategy_b(detail, mappings):
    """
    Mapping bertipe 'Cost center' -> pisahkan per CC dari API detail.
    Mapping bertipe 'Aggregate' -> sum semua, output 1 row.
    Surabaya version: match account_name pakai match_account_name field.
    Pre-init semua mapping dengan amount=0 (agar selalu muncul).
    """
    gl_data = {}
    for m in mappings:
        gl_data[m['mapping_key']] = {'mapping': m, 'rows': []}

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = cfg.get('account_number', '')
        account_name = cfg.get('account_name', '').strip()
        account_name_lower = account_name.lower()

        # Match dengan strategi 2 (account_number + match_account_name)
        matched_key = match_by_match_account_name(mappings, account_number, account_name_lower)
        if not matched_key:
            # Fallback: keywords
            matched_key = match_by_account_number_and_keywords(mappings, account_number, account_name_lower)
        if not matched_key:
            continue

        mapping = gl_data[matched_key]['mapping']

        if mapping['account_type'] == 'Cost center':
            for d in report_item.get('details', []):
                amount = parse_amount(d.get('amount', 0))
                cc_code = d.get('cost_center', '') or d.get('cost_center_code', '')
                cc_name = d.get('cost_center_name', '') or d.get('cost_center', '')
                gl_data[matched_key]['rows'].append({
                    'amount': amount,
                    'cost_center_code': str(cc_code) if cc_code else '',
                    'cost_center_name': cc_name,
                })
        else:  # Aggregate
            total = 0
            for d in report_item.get('details', []):
                total += parse_amount(d.get('amount', 0))
            if gl_data[matched_key]['rows']:
                gl_data[matched_key]['rows'][0]['amount'] += total
            else:
                gl_data[matched_key]['rows'].append({
                    'amount': total,
                    'cost_center_code': '',
                    'cost_center_name': '',
                })

    # Build entries
    entries = []
    for m in mappings:
        key = m['mapping_key']
        rows = gl_data[key]['rows']

        if m['account_type'] == 'Aggregate':
            # Pre-init kalau gak ada data
            if not rows:
                rows = [{'amount': 0, 'cost_center_code': '', 'cost_center_name': ''}]
            for r in rows:
                entries.append({
                    'account_number': m['account_number'],
                    'transaction_value': m['transaction_value'],
                    'amount': r['amount'],
                    'cost_center': '',
                    'profit_center': m['profit_center'] if m['use_profit_center'] else '',
                })
        else:  # Cost center
            for r in rows:
                entries.append({
                    'account_number': m['account_number'],
                    'transaction_value': m['transaction_value'],
                    'amount': r['amount'],
                    'cost_center': r['cost_center_code'],
                    'profit_center': '',
                })

    return entries


# ============================================================
# STRATEGY D: Auto-detect D/C (no mapping)
# ============================================================
def extract_strategy_d(detail, strategy_d_config):
    """
    Tanpa mapping eksplisit.
    D/C ditentukan dari:
    1. Keyword di account_name (e.g., 'pengembalian' -> Debit)
    2. Whitelist account_number -> Debit
    3. Default -> Credit

    account_type ditentukan dari:
    - Ada cost_center di details -> Cost center
    - Semua kosong -> Aggregate
    """
    debit_accounts = set(strategy_d_config.get('debit_accounts', []))
    debit_keywords = [kw.lower() for kw in strategy_d_config.get('debit_keywords', [])]
    default_dc = strategy_d_config.get('default_dc', 'Credit')

    entries = []

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = str(cfg.get('account_number', ''))
        account_name = cfg.get('account_name', '').strip()
        account_name_lower = account_name.lower()
        details = report_item.get('details', [])

        if not account_number or not details:
            continue

        # Determine D/C
        trans_value = default_dc
        # Priority 1: keyword di account_name
        for kw in debit_keywords:
            if kw in account_name_lower:
                trans_value = 'Debit'
                break
        # Priority 2: whitelist account_number
        if trans_value == default_dc and account_number in debit_accounts:
            trans_value = 'Debit'

        # Determine account_type berdasarkan ada/tidaknya cost_center di details
        has_cost_center = False
        for d in details:
            cc = d.get('cost_center', '')
            cc_id = d.get('cost_center_id', '')
            if cc and cc != '-' and cc_id and cc_id != '-':
                has_cost_center = True
                break
        account_type = 'Cost center' if has_cost_center else 'Aggregate'

        if account_type == 'Aggregate':
            total = 0
            for d in details:
                total += parse_amount(d.get('amount', 0))
            entries.append({
                'account_number': account_number,
                'transaction_value': trans_value,
                'amount': total,
                'cost_center': '',
                'profit_center': '200301',
            })
        else:  # Cost center
            for d in details:
                amount = parse_amount(d.get('amount', 0))
                cc = d.get('cost_center', '')
                if cc == '-':
                    cc = ''
                entries.append({
                    'account_number': account_number,
                    'transaction_value': trans_value,
                    'amount': amount,
                    'cost_center': str(cc) if cc else '',
                    'profit_center': '',
                })

    return entries


# ============================================================
# STRATEGY E: Mapping with preserved API order (KMI1)
# ============================================================
def extract_strategy_e(detail, mappings):
    """
    Sama dengan strategy B, TAPI urutan output preserve dari API.
    Tidak di-reorder berdasarkan order_index mapping.
    Per detail item di-emit langsung (preserve sequence dari API).
    """
    # Build mapping lookup
    mapping_by_key = {m['mapping_key']: m for m in mappings}

    entries = []  # Output entries dalam urutan API

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = cfg.get('account_number', '')
        account_name = cfg.get('account_name', '').strip()
        account_name_lower = account_name.lower()

        matched_key = match_by_account_number_and_keywords(mappings, account_number, account_name_lower)
        if not matched_key:
            continue

        mapping = mapping_by_key[matched_key]
        details = report_item.get('details', [])

        if mapping['account_type'] == 'Aggregate':
            # Setiap detail = 1 entry (untuk preserve count di API)
            for d in details:
                amount = parse_amount(d.get('amount', 0))
                entries.append({
                    'account_number': mapping['account_number'],
                    'transaction_value': mapping['transaction_value'],
                    'amount': amount,
                    'cost_center': '',
                    'profit_center': '200301',
                })
        else:  # Cost center
            for d in details:
                amount = parse_amount(d.get('amount', 0))
                # PENTING: pakai field 'cost_center' (kode numerik), bukan 'cost_center_id'
                cc_code = d.get('cost_center', '')
                if cc_code == '-':
                    cc_code = ''
                entries.append({
                    'account_number': mapping['account_number'],
                    'transaction_value': mapping['transaction_value'],
                    'amount': amount,
                    'cost_center': str(cc_code) if cc_code else '',
                    'profit_center': '',
                })

    return entries


# ============================================================
# STRATEGY F: Per-detail include-zero (Pembantu KMI)
# ============================================================
def extract_strategy_f(detail, mappings):
    """
    Setiap detail item = 1 row SAP, TERMASUK amount=0.
    Match by account_number exact.
    """
    mapping_by_account = {m['account_number']: m for m in mappings}

    entries = []

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = cfg.get('account_number', '')

        if account_number not in mapping_by_account:
            continue

        mapping = mapping_by_account[account_number]

        for d in report_item.get('details', []):
            amount = parse_amount(d.get('amount', 0))
            cc = d.get('cost_center', '')
            if cc == '-':
                cc = ''

            if mapping['account_type'] == 'Cost center':
                cost_center_display = str(cc) if cc else ''
                profit_center = ''
            else:  # Aggregate
                cost_center_display = ''
                profit_center = '200301'

            entries.append({
                'account_number': account_number,
                'transaction_value': mapping['transaction_value'],
                'amount': amount,
                'cost_center': cost_center_display,
                'profit_center': profit_center,
            })

    # Sort: Debit dulu, Credit kemudian
    entries.sort(key=lambda e: (0 if e['transaction_value'] == 'Debit' else 1, e['account_number']))

    return entries


# ============================================================
# CONVERT TO SAP FORMAT (sama dengan Semarang)
# ============================================================
def convert_to_sap_format(entries, config, year, month):
    if not entries:
        return pd.DataFrame()

    last_day = 31 if month in [1, 3, 5, 7, 8, 10, 12] else (30 if month in [4, 6, 9, 11] else
              (29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28))

    doc_date = f"{last_day:02d}.{month:02d}.{year}"
    doc_header = config['doc_header_template'] \
        .replace('{MONTH}', MONTH_NAMES[month]) \
        .replace('{YEAR}', str(year))

    sap_entries = []
    for entry in entries:
        posting_key = 40 if entry['transaction_value'] == 'Debit' else 50
        sap_entries.append({
            'Document Date': doc_date,
            'Posting Date': doc_date,
            'Doc. Type': SAP_BASE_SETTINGS['doc_type'],
            'Company Code': config.get('company_code', 'KMI'),
            'Curr': SAP_BASE_SETTINGS['currency'],
            'Reference': SAP_BASE_SETTINGS['reference'],
            'Doc. Header': doc_header,
            'PstKy': int(posting_key),
            'Account': int(entry['account_number']),
            'Sp.G/L': '',
            'Amount': int(entry['amount']) if entry['amount'] else 0,
            'Due On': '',
            'Tax Code': '',
            'Value Date': '',
            'Cost Center': entry.get('cost_center', ''),
            'Profit Center': entry.get('profit_center', ''),
            'Assignment': '',
            'Text': '',
            'Reason Code': '',
            'House Bank': '',
        })

    return pd.DataFrame(sap_entries)


# ============================================================
# EXCEL WRITER (sama dengan Semarang)
# ============================================================
def save_excel(df, filename):
    if os.path.exists(filename):
        os.remove(filename)
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='GL_Result', index=False)
        ws = writer.sheets['GL_Result']
        _format_sheet(ws, df)


def _format_sheet(ws, df):
    header_font = Font(bold=True, size=11)
    data_font = Font(bold=False, size=10)
    header_fill = PatternFill(start_color='A5A5A5', end_color='A5A5A5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if df.empty:
        return

    num_cols = len(df.columns)
    num_rows = len(df) + 1

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

            col_name = df.columns[col - 1]
            value = df.iloc[row - 2, col - 1]

            if col_name in ['Amount', 'PstKy', 'Account']:
                cell.value = int(value) if pd.notna(value) else 0
                cell.number_format = '0'
            elif col_name in ['Cost Center', 'Profit Center']:
                if pd.notna(value) and str(value).strip():
                    sv = str(value).strip()
                    if sv.isdigit():
                        cell.value = int(sv)
                        cell.number_format = '0'
                    else:
                        cell.value = sv
                        cell.number_format = '@'
                else:
                    cell.value = ''
                    cell.number_format = '@'
            else:
                cell.value = str(value) if pd.notna(value) and value != '' else ''
                cell.number_format = '@'

    col_widths = {'A': 12, 'B': 12, 'C': 10, 'D': 12, 'E': 6, 'F': 10,
                  'G': 30, 'H': 8, 'I': 12, 'J': 8, 'K': 12, 'L': 8,
                  'M': 8, 'N': 10, 'O': 16, 'P': 12, 'Q': 10, 'R': 8, 'S': 10, 'T': 10}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_border


# ============================================================
# MAIN PIPELINE
# ============================================================
def execute_extraction(config):
    year = config['year']
    month = config['month']
    client_id = config['talenta']['client_id']
    client_secret = config['talenta']['client_secret']
    ledger_code = config['ledger_code']
    strategy = config['extraction_strategy']
    mappings = config.get('mappings', [])
    strategy_d_config = config.get('strategy_d_config', {})

    # 1. Fetch report history
    reports = get_ledger_report_history(year, month, client_id, client_secret)
    if not reports:
        return {'status': 'failed', 'error': 'No reports found in Talenta history'}

    # 2. Filter by ledger_code
    selected = filter_reports_by_ledger_code(reports, ledger_code)
    if not selected:
        return {'status': 'failed', 'error': f"Report with ledger_code '{ledger_code}' not found in latest history"}

    report_id = selected.get('id')

    # 3. Fetch detail (single or multi-try ledger_id)
    ledger_ids = config.get('ledger_id_list', [900])
    best_detail = None
    best_count = 0

    for lid in ledger_ids:
        d = get_report_detail(report_id, lid, client_id, client_secret)
        if d and 'reports' in d:
            count = sum(len(r.get('details', [])) for r in d.get('reports', []))
            if count > best_count:
                best_detail = d
                best_count = count
        time.sleep(0.1)

    if not best_detail:
        return {'status': 'failed', 'error': f'No detail data found for report_id {report_id}'}

    # 4. Run strategy
    if strategy == 'B':
        entries = extract_strategy_b(best_detail, mappings)
    elif strategy == 'D':
        if not strategy_d_config:
            return {'status': 'failed', 'error': 'Strategy D requires strategy_d_config'}
        entries = extract_strategy_d(best_detail, strategy_d_config)
    elif strategy == 'E':
        entries = extract_strategy_e(best_detail, mappings)
    elif strategy == 'F':
        entries = extract_strategy_f(best_detail, mappings)
    else:
        return {'status': 'failed', 'error': f'Strategy {strategy} not supported in Surabaya service'}

    if not entries:
        return {'status': 'failed', 'error': 'Strategy execution returned no entries'}

    # 5. Convert to SAP format
    df = convert_to_sap_format(entries, config, year, month)

    # 6. Save Excel
    filename = config['output_filename_template'] \
        .replace('{YEAR}', str(year)) \
        .replace('{MONTH}', f"{month:02d}")
    output_path = os.path.join(OUTPUT_BASE_DIR, filename)
    save_excel(df, output_path)

    # 7. Compute totals
    total_debit = int(df[df['PstKy'] == 40]['Amount'].sum())
    total_credit = int(df[df['PstKy'] == 50]['Amount'].sum())

    return {
        'status': 'success',
        'output_file': output_path,
        'output_filename': filename,
        'total_records': len(df),
        'total_debit': total_debit,
        'total_credit': total_credit,
        'difference': total_debit - total_credit,
        'report_id': report_id,
        'ledger_name': selected.get('ledger_name'),
        'created_at': selected.get('created_at'),
    }


# ============================================================
# FLASK APP
# ============================================================
app = Flask(__name__)


@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'service': SERVICE_NAME,
        'port': SERVICE_PORT,
        'strategies_supported': ['B', 'D', 'E', 'F'],
        'timestamp': datetime.now().isoformat(),
    })


@app.route('/run', methods=['POST'])
def run():
    try:
        config = request.get_json(force=True)

        if not config:
            return jsonify({'status': 'failed', 'error': 'No config provided'}), 400

        required_fields = ['entity_code', 'ledger_code', 'year', 'month',
                           'extraction_strategy', 'talenta']
        missing = [f for f in required_fields if f not in config]
        if missing:
            return jsonify({'status': 'failed', 'error': f'Missing fields: {missing}'}), 400

        result = execute_extraction(config)
        status_code = 200 if result['status'] == 'success' else 500
        return jsonify(result), status_code

    except Exception as e:
        return jsonify({
            'status': 'failed',
            'error': str(e),
            'trace': traceback.format_exc(),
        }), 500


if __name__ == '__main__':
    os.makedirs(OUTPUT_BASE_DIR, exist_ok=True)
    print("=" * 60)
    print(f"  {SERVICE_NAME}")
    print(f"  Listening on http://127.0.0.1:{SERVICE_PORT}")
    print(f"  Output dir: {OUTPUT_BASE_DIR}")
    print(f"  Strategies: B, D, E, F")
    print("=" * 60)
    app.run(host='127.0.0.1', port=SERVICE_PORT, debug=False)