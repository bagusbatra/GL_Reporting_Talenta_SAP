"""
GL Reporting Python Service - SEMARANG (Port 8091)
====================================================
Service untuk fetch data Talenta + generate Excel SAP format.
Support 3 extraction strategy: A, B, C

Endpoint:
    POST /run    -> jalankan extraction
    GET  /health -> health check

Dipanggil oleh Laravel via HTTP request.
"""

import hmac
import hashlib
import base64
import json
import os
import time
import traceback
from datetime import datetime
from urllib.parse import urlparse

import requests
import pandas as pd
from flask import Flask, request, jsonify
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ============================================================
# CONFIG
# ============================================================
SERVICE_PORT = 8091
SERVICE_NAME = "GL Reporting Service - SEMARANG"
OUTPUT_BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'storage', 'app', 'gl_outputs')

SAP_BASE_SETTINGS = {
    'doc_type': 'KA',
    'currency': 'IDR',
    'reference': 'TALENTA',
}

MONTH_NAMES = ["", "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
               "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"]

# ============================================================
# TALENTA API CLIENT
# ============================================================
def api_request(url, client_id, client_secret):
    """Authenticated request ke Talenta API pakai HMAC SHA-256"""
    try:
        parsed = urlparse(url)
        request_line = f"GET {parsed.path}?{parsed.query} HTTP/1.1"
        date_string = datetime.utcnow().strftime('%a, %d %b %Y %H:%M:%S GMT')
        signing_string = f"date: {date_string}\n{request_line}"
        digest = hmac.new(client_secret.encode(), signing_string.encode(), hashlib.sha256).digest()
        signature = base64.b64encode(digest).decode()
        headers = {
            "Authorization": f'hmac username="{client_id}", algorithm="hmac-sha256", headers="date request-line", signature="{signature}"',
            "Date": date_string
        }
        response = requests.get(url, headers=headers, timeout=30)
        if response.status_code == 200:
            return response.json()
        elif response.status_code == 429:
            time.sleep(2)
            return api_request(url, client_id, client_secret)
        else:
            return None
    except Exception as e:
        print(f"API request error: {e}")
        return None


def get_ledger_report_history(year, month, client_id, client_secret):
    url = f"https://api.mekari.com/v2/talenta/v3/ledger/report/history?year={year}&month={month}&page=1&limit=100"
    data = api_request(url, client_id, client_secret)
    if data and 'data' in data:
        return data['data'].get('ledger_results', [])
    return []


def filter_reports_by_ledger_code(reports, target_ledger_code):
    matching = []
    for r in reports:
        code = r.get('ledger_code', '').strip()
        name = r.get('ledger_name', '').strip()
        if (code == target_ledger_code or name == target_ledger_code or
                target_ledger_code in code or target_ledger_code in name):
            matching.append(r)
    if not matching:
        return None
    matching.sort(key=lambda r: r.get('created_at', ''), reverse=True)
    return matching[0]


def get_report_detail(report_id, ledger_id, client_id, client_secret):
    url = f"https://api.mekari.com/v2/talenta/v3/ledger/report/{report_id}?ledger_id={ledger_id}"
    data = api_request(url, client_id, client_secret)
    if data and 'data' in data:
        return data['data'].get('ledger_result', {})
    return {}


def parse_amount(val):
    if isinstance(val, str):
        try:
            return float(val.replace(',', ''))
        except:
            return 0
    return val or 0


# ============================================================
# STRATEGY A: Aggregate-only with fixed CC
# ============================================================
def extract_strategy_a(detail, mappings):
    """
    Setiap mapping key = 1 Aggregate entry.
    Sum semua details per matched account_name, output 1 row.
    Cost Center fixed dari mapping, atau profit_center 200301.
    Pre-init semua mapping dengan amount=0 agar selalu muncul.
    """
    # Pre-init semua mapping
    gl_data = {}
    for m in mappings:
        gl_data[m['mapping_key']] = {
            'mapping': m,
            'total_amount': 0,
        }

    # Match report items
    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_name = cfg.get('account_name', '').strip()
        account_name_lower = account_name.lower()

        # Cari mapping yang cocok via match_keywords
        matched_key = match_by_keywords(mappings, account_name_lower)
        if not matched_key:
            continue

        # Sum amount
        total = 0
        for d in report_item.get('details', []):
            total += parse_amount(d.get('amount', 0))

        gl_data[matched_key]['total_amount'] += total

    # Build entries (sorted by order_index in mappings input)
    entries = []
    for m in mappings:
        key = m['mapping_key']
        total = gl_data[key]['total_amount']
        entry = {
            'account_number': m['account_number'],
            'transaction_value': m['transaction_value'],
            'amount': total,
            'cost_center': m['cost_center'] if not m['use_profit_center'] else '',
            'profit_center': m['profit_center'] if m['use_profit_center'] else '',
        }
        entries.append(entry)

    return entries


def match_by_keywords(mappings, account_name_lower):
    """Cari mapping yang match berdasarkan match_keywords (longest match first)"""
    candidates = []
    for m in mappings:
        keywords = m.get('match_keywords') or []
        if not keywords:
            # Fallback: match by account_name dari mapping
            if m.get('account_name', '').lower() in account_name_lower:
                candidates.append((m['mapping_key'], len(m['account_name'])))
            continue
        for kw in keywords:
            if kw.lower() in account_name_lower:
                candidates.append((m['mapping_key'], len(kw)))
                break

    if not candidates:
        return None
    # Pilih yang paling spesifik (keyword paling panjang)
    candidates.sort(key=lambda x: x[1], reverse=True)
    return candidates[0][0]


# ============================================================
# STRATEGY B: Hybrid Cost Center + Aggregate
# ============================================================
def extract_strategy_b(detail, mappings):
    """
    Mapping bertipe 'Cost center' -> pisahkan per CC dari API detail.
    Mapping bertipe 'Aggregate' -> sum semua, output 1 row.
    """
    # Pre-init aggregate dengan amount=0; cost center kosong
    gl_data = {}
    for m in mappings:
        gl_data[m['mapping_key']] = {
            'mapping': m,
            'rows': [],  # list of {amount, cost_center_code, cost_center_name}
        }

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = cfg.get('account_number', '')
        account_name = cfg.get('account_name', '').strip()
        account_name_lower = account_name.lower()

        # Match by account_number first, then keywords
        matched_key = match_by_account_number_and_keywords(
            mappings, account_number, account_name_lower
        )
        if not matched_key:
            continue

        mapping = gl_data[matched_key]['mapping']

        if mapping['account_type'] == 'Cost center':
            # Pisahkan per cost center dari detail
            for d in report_item.get('details', []):
                amount = parse_amount(d.get('amount', 0))
                cc_code = d.get('cost_center_code', '') or d.get('cost_center', '') or d.get('costcenter_code', '')
                cc_name = d.get('cost_center_name', '') or d.get('cost_center', '')
                gl_data[matched_key]['rows'].append({
                    'amount': amount,
                    'cost_center_code': cc_code,
                    'cost_center_name': cc_name,
                })
        else:  # Aggregate
            total = 0
            for d in report_item.get('details', []):
                total += parse_amount(d.get('amount', 0))
            # Append/update aggregate row
            if gl_data[matched_key]['rows']:
                gl_data[matched_key]['rows'][0]['amount'] += total
            else:
                gl_data[matched_key]['rows'].append({
                    'amount': total,
                    'cost_center_code': '',
                    'cost_center_name': '',
                })

    # Build entries
    entries = []
    for m in mappings:
        key = m['mapping_key']
        rows = gl_data[key]['rows']

        if m['account_type'] == 'Aggregate':
            # Pre-init: kalau gak ada data, tetep output amount=0
            if not rows:
                rows = [{'amount': 0, 'cost_center_code': '', 'cost_center_name': ''}]
            for r in rows:
                entries.append({
                    'account_number': m['account_number'],
                    'transaction_value': m['transaction_value'],
                    'amount': r['amount'],
                    'cost_center': '',
                    'profit_center': m['profit_center'] if m['use_profit_center'] else '',
                })
        else:  # Cost center
            # Kalau gak ada data, skip (sesuai logic original)
            for r in rows:
                entries.append({
                    'account_number': m['account_number'],
                    'transaction_value': m['transaction_value'],
                    'amount': r['amount'],
                    'cost_center': r['cost_center_code'],
                    'profit_center': '',
                })

    return entries


def match_by_account_number_and_keywords(mappings, account_number, account_name_lower):
    """
    Match by:
    1. exact account_number (kalau cuma 1 mapping dengan account_number ini)
    2. account_number + keyword (kalau multi-variant)
    """
    # Kumpulkan mapping dengan account_number ini
    candidates = [m for m in mappings if m['account_number'] == account_number]

    if not candidates:
        return None

    if len(candidates) == 1:
        return candidates[0]['mapping_key']

    # Multi-variant: pilih berdasarkan keyword match
    best_match = None
    best_kw_len = 0
    for m in candidates:
        keywords = m.get('match_keywords') or []
        for kw in keywords:
            if kw.lower() in account_name_lower and len(kw) > best_kw_len:
                best_match = m['mapping_key']
                best_kw_len = len(kw)

    return best_match if best_match else candidates[0]['mapping_key']


# ============================================================
# STRATEGY C: Individual (per-detail, no aggregation)
# ============================================================
def extract_strategy_c(detail, mappings):
    """
    Setiap detail item dari API = 1 row di output.
    Match by account_number exact (no aggregation).
    """
    # Build account_number -> mapping lookup
    mapping_by_account = {}
    for m in mappings:
        mapping_by_account[m['account_number']] = m

    entries = []

    for report_item in detail.get('reports', []):
        cfg = report_item.get('config', {})
        account_number = cfg.get('account_number', '')

        if account_number not in mapping_by_account:
            continue

        mapping = mapping_by_account[account_number]

        for d in report_item.get('details', []):
            amount = parse_amount(d.get('amount', 0))
            cc = d.get('cost_center', '')
            if cc == '-':
                cc = ''

            if cc:
                cost_center_display = cc
                profit_center = ''
            else:
                cost_center_display = ''
                profit_center = '200301'

            entries.append({
                'account_number': account_number,
                'transaction_value': mapping['transaction_value'],
                'amount': amount,
                'cost_center': cost_center_display,
                'profit_center': profit_center,
            })

    # Sort: Debit dulu, Credit kemudian, by account_number
    entries.sort(key=lambda e: (50 if e['transaction_value'] == 'Credit' else 40, e['account_number']))

    return entries


# ============================================================
# CONVERT TO SAP FORMAT (20 KOLOM)
# ============================================================
def convert_to_sap_format(entries, config, year, month):
    if not entries:
        return pd.DataFrame()

    last_day = 31 if month in [1, 3, 5, 7, 8, 10, 12] else (30 if month in [4, 6, 9, 11] else
              (29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28))

    doc_date = f"{last_day:02d}.{month:02d}.{year}"
    doc_header = config['doc_header_template'] \
        .replace('{MONTH}', MONTH_NAMES[month]) \
        .replace('{YEAR}', str(year))

    sap_entries = []
    for entry in entries:
        posting_key = 40 if entry['transaction_value'] == 'Debit' else 50
        sap_entries.append({
            'Document Date': doc_date,
            'Posting Date': doc_date,
            'Doc. Type': SAP_BASE_SETTINGS['doc_type'],
            'Company Code': config.get('company_code', 'KMI'),
            'Curr': SAP_BASE_SETTINGS['currency'],
            'Reference': SAP_BASE_SETTINGS['reference'],
            'Doc. Header': doc_header,
            'PstKy': int(posting_key),
            'Account': int(entry['account_number']),
            'Sp.G/L': '',
            'Amount': int(entry['amount']) if entry['amount'] else 0,
            'Due On': '',
            'Tax Code': '',
            'Value Date': '',
            'Cost Center': entry.get('cost_center', ''),
            'Profit Center': entry.get('profit_center', ''),
            'Assignment': '',
            'Text': '',
            'Reason Code': '',
            'House Bank': '',
        })

    return pd.DataFrame(sap_entries)


# ============================================================
# EXCEL WRITER
# ============================================================
def save_excel(df, filename):
    if os.path.exists(filename):
        os.remove(filename)
    os.makedirs(os.path.dirname(filename), exist_ok=True)

    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='GL_Result', index=False)
        ws = writer.sheets['GL_Result']
        _format_sheet(ws, df)


def _format_sheet(ws, df):
    header_font = Font(bold=True, size=11)
    data_font = Font(bold=False, size=10)
    header_fill = PatternFill(start_color='A5A5A5', end_color='A5A5A5', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    if df.empty:
        return

    num_cols = len(df.columns)
    num_rows = len(df) + 1

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for row in range(2, num_rows + 1):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.alignment = Alignment(horizontal='left', vertical='center')

            col_name = df.columns[col - 1]
            value = df.iloc[row - 2, col - 1]

            if col_name in ['Amount', 'PstKy', 'Account']:
                cell.value = int(value) if pd.notna(value) else 0
                cell.number_format = '0'
            elif col_name in ['Cost Center', 'Profit Center']:
                if pd.notna(value) and str(value).strip():
                    sv = str(value).strip()
                    if sv.isdigit():
                        cell.value = int(sv)
                        cell.number_format = '0'
                    else:
                        cell.value = sv
                        cell.number_format = '@'
                else:
                    cell.value = ''
                    cell.number_format = '@'
            else:
                cell.value = str(value) if pd.notna(value) and value != '' else ''
                cell.number_format = '@'

    col_widths = {'A': 12, 'B': 12, 'C': 10, 'D': 12, 'E': 6, 'F': 10,
                  'G': 30, 'H': 8, 'I': 12, 'J': 8, 'K': 12, 'L': 8,
                  'M': 8, 'N': 10, 'O': 16, 'P': 12, 'Q': 10, 'R': 8, 'S': 10, 'T': 10}
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_border


# ============================================================
# MAIN PIPELINE
# ============================================================
def execute_extraction(config):
    """
    config (dict) berisi:
    - entity_code, entity_name, ledger_code, branch_id
    - ledger_id_strategy, ledger_id_list
    - doc_header_template, output_filename_template
    - extraction_strategy: 'A', 'B', atau 'C'
    - company_code
    - year, month
    - mappings: list of mapping dict
    - talenta: {client_id, client_secret}
    """
    year = config['year']
    month = config['month']
    client_id = config['talenta']['client_id']
    client_secret = config['talenta']['client_secret']
    ledger_code = config['ledger_code']
    strategy = config['extraction_strategy']
    mappings = config.get('mappings', [])

    # 1. Fetch report history
    reports = get_ledger_report_history(year, month, client_id, client_secret)
    if not reports:
        return {'status': 'failed', 'error': 'No reports found in Talenta history'}

    # 2. Filter by ledger_code
    selected = filter_reports_by_ledger_code(reports, ledger_code)
    if not selected:
        return {'status': 'failed', 'error': f"Report with ledger_code '{ledger_code}' not found in latest history"}

    report_id = selected.get('id')

    # 3. Fetch detail (single or multi-try ledger_id)
    ledger_ids = config.get('ledger_id_list', [900])
    best_detail = None
    best_count = 0

    for lid in ledger_ids:
        detail = get_report_detail(report_id, lid, client_id, client_secret)
        if detail and 'reports' in detail:
            count = sum(len(r.get('details', [])) for r in detail.get('reports', []))
            if count > best_count:
                best_detail = detail
                best_count = count
        time.sleep(0.1)

    if not best_detail:
        return {'status': 'failed', 'error': f'No detail data found for report_id {report_id}'}

    # 4. Run strategy
    if strategy == 'A':
        entries = extract_strategy_a(best_detail, mappings)
    elif strategy == 'B':
        entries = extract_strategy_b(best_detail, mappings)
    elif strategy == 'C':
        entries = extract_strategy_c(best_detail, mappings)
    else:
        return {'status': 'failed', 'error': f'Strategy {strategy} not supported in Semarang service'}

    if not entries:
        return {'status': 'failed', 'error': 'Strategy execution returned no entries'}

    # 5. Convert to SAP format
    df = convert_to_sap_format(entries, config, year, month)

    # 6. Save Excel
    filename = config['output_filename_template'] \
        .replace('{YEAR}', str(year)) \
        .replace('{MONTH}', f"{month:02d}")
    output_path = os.path.join(OUTPUT_BASE_DIR, filename)
    save_excel(df, output_path)

    # 7. Compute totals
    total_debit = int(df[df['PstKy'] == 40]['Amount'].sum())
    total_credit = int(df[df['PstKy'] == 50]['Amount'].sum())

    return {
        'status': 'success',
        'output_file': output_path,
        'output_filename': filename,
        'total_records': len(df),
        'total_debit': total_debit,
        'total_credit': total_credit,
        'difference': total_debit - total_credit,
        'report_id': report_id,
        'ledger_name': selected.get('ledger_name'),
        'created_at': selected.get('created_at'),
    }


# ============================================================
# FLASK APP
# ============================================================
app = Flask(__name__)


@app.route('/health', methods=['GET'])
def health():
    return jsonify({
        'status': 'ok',
        'service': SERVICE_NAME,
        'port': SERVICE_PORT,
        'strategies_supported': ['A', 'B', 'C'],
        'timestamp': datetime.now().isoformat(),
    })


@app.route('/run', methods=['POST'])
def run():
    try:
        config = request.get_json(force=True)

        if not config:
            return jsonify({'status': 'failed', 'error': 'No config provided'}), 400

        required_fields = ['entity_code', 'ledger_code', 'year', 'month',
                           'extraction_strategy', 'talenta']
        missing = [f for f in required_fields if f not in config]
        if missing:
            return jsonify({'status': 'failed', 'error': f'Missing fields: {missing}'}), 400

        result = execute_extraction(config)
        status_code = 200 if result['status'] == 'success' else 500
        return jsonify(result), status_code

    except Exception as e:
        return jsonify({
            'status': 'failed',
            'error': str(e),
            'trace': traceback.format_exc(),
        }), 500


if __name__ == '__main__':
    os.makedirs(OUTPUT_BASE_DIR, exist_ok=True)
    print("=" * 60)
    print(f"  {SERVICE_NAME}")
    print(f"  Listening on http://127.0.0.1:{SERVICE_PORT}")
    print(f"  Output dir: {OUTPUT_BASE_DIR}")
    print(f"  Strategies: A, B, C")
    print("=" * 60)
    app.run(host='127.0.0.1', port=SERVICE_PORT, debug=False)